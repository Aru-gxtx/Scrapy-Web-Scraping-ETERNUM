from pathlib import Path

from openpyxl import load_workbook


EXCEL_FILE = Path("sources/ETERNUM.xlsx")
START_COLUMN = 6  # F


def is_missing(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def analyze_sheet(ws, start_column=START_COLUMN):
    max_row = ws.max_row
    max_column = ws.max_column

    if max_row < 2 or max_column < start_column:
        return {
            "data_rows": 0,
            "checked_columns": 0,
            "total_cells": 0,
            "missing_cells": 0,
            "filled_cells": 0,
            "missing_ratio": 0.0,
            "filled_ratio": 0.0,
            "rows_with_missing": 0,
            "rows_fully_filled": 0,
            "row_missing_ratio": 0.0,
            "row_filled_ratio": 0.0,
        }

    data_rows = max_row - 1
    checked_columns = max_column - start_column + 1
    total_cells = data_rows * checked_columns

    missing_cells = 0
    rows_with_missing = 0

    for row_idx in range(2, max_row + 1):
        row_missing = 0
        for col_idx in range(start_column, max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if is_missing(value):
                row_missing += 1

        if row_missing > 0:
            rows_with_missing += 1
        missing_cells += row_missing

    filled_cells = total_cells - missing_cells
    rows_fully_filled = data_rows - rows_with_missing

    missing_ratio = (missing_cells / total_cells * 100) if total_cells else 0.0
    filled_ratio = 100.0 - missing_ratio if total_cells else 0.0
    row_missing_ratio = (rows_with_missing / data_rows * 100) if data_rows else 0.0
    row_filled_ratio = 100.0 - row_missing_ratio if data_rows else 0.0

    return {
        "data_rows": data_rows,
        "checked_columns": checked_columns,
        "total_cells": total_cells,
        "missing_cells": missing_cells,
        "filled_cells": filled_cells,
        "missing_ratio": missing_ratio,
        "filled_ratio": filled_ratio,
        "rows_with_missing": rows_with_missing,
        "rows_fully_filled": rows_fully_filled,
        "row_missing_ratio": row_missing_ratio,
        "row_filled_ratio": row_filled_ratio,
    }


def main():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, keep_links=False)
    ws = wb.active

    stats = analyze_sheet(ws, START_COLUMN)

    print("=" * 70)
    print("ETERNUM Excel Missing Data Report (from column F onward)")
    print("=" * 70)
    print(f"File:                     {EXCEL_FILE}")
    print(f"Sheet:                    {ws.title}")
    print(f"Data rows checked:        {stats['data_rows']}")
    print(f"Columns checked (F->end): {stats['checked_columns']}")
    print("-" * 70)
    print(f"Total cells checked:      {stats['total_cells']}")
    print(f"Missing cells:            {stats['missing_cells']}")
    print(f"Filled cells:             {stats['filled_cells']}")
    print(f"Missing ratio:            {stats['missing_ratio']:.2f}%")
    print(f"Filled ratio:             {stats['filled_ratio']:.2f}%")
    print("-" * 70)
    print(f"Rows with missing data:   {stats['rows_with_missing']}")
    print(f"Rows fully filled:        {stats['rows_fully_filled']}")
    print(f"Row missing ratio:        {stats['row_missing_ratio']:.2f}%")
    print(f"Row filled ratio:         {stats['row_filled_ratio']:.2f}%")
    print("=" * 70)


if __name__ == "__main__":
    main()