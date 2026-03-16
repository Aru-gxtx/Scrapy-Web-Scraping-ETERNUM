import json
import re
from openpyxl import load_workbook
from pathlib import Path

# File paths
JSON_FILES = [
    r"eternum\horecaservise_v0.1.json",
    r"eternum\yourroyalhouse.json",
    r"eternum\zakaz.json",
    r"eternum\tomgast.json",
    r"eternum\entero.json",
    r"eternum\liberty.json",
]
EXCEL_FILE = r"sources\ETERNUM.xlsx"

def extract_catalog_number(title):
    match = re.match(r'^(\d+-\d+)', title.strip())
    return match.group(1) if match else None


def normalize_catalog_number(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def get_catalog_number(item):
    sku = normalize_catalog_number(item.get("sku", ""))
    if sku:
        return sku

    title = item.get("title", "")
    extracted = extract_catalog_number(title)
    if extracted:
        return extracted

    return normalize_catalog_number(item.get("product_key", ""))

def load_json_data(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_all_json_data(json_paths):
    all_items = []
    loaded_sources = []

    for path in json_paths:
        items = load_json_data(path)
        if not isinstance(items, list):
            raise ValueError(f"JSON root must be a list in: {path}")
        all_items.extend(items)
        loaded_sources.append((path, len(items)))

    return all_items, loaded_sources

def load_excel():
    if Path(EXCEL_FILE).exists():
        wb = load_workbook(EXCEL_FILE, keep_links=False)
        ws = wb.active
        print(f"[OK] Loaded existing Excel file: {EXCEL_FILE}")
    else:
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    
    return wb, ws


def pick_best_item(items):
    for item in items:
        image_url = item.get('image_url') or item.get('listing_image_url') or ''
        if isinstance(image_url, str) and image_url.startswith('http'):
            return item
    return items[0] if items else None


def get_preferred_price(item):
    return (
        item.get('price')
        or item.get('price_incl_tax')
        or item.get('price_excl_tax')
        or ''
    )


def get_preferred_image(item):
    return item.get('image_url') or item.get('listing_image_url') or ''


def get_detail_image(item):
    detail_image = item.get('detail_image_url')
    if detail_image:
        return detail_image

    image_urls = item.get('image_urls')
    if isinstance(image_urls, list) and image_urls:
        first_image = image_urls[0]
        return first_image if isinstance(first_image, str) else ''

    return item.get('listing_image_url') or item.get('image_url') or ''


def get_product_url(item):
    return item.get('url') or item.get('product_page') or ''

def populate_excel(wb, ws, json_data):
    catalog_lookup = {}
    for item in json_data:
        catalog = get_catalog_number(item)
        if catalog:
            catalog_lookup.setdefault(catalog, []).append(item)
    
    matched_count = 0
    unmatched_count = 0
    
    # Column positions (1-indexed) - ONLY populate F onwards
    COL_CATALOG = 2      # B: Mfr Catalog No. (reference only, do NOT change)
    COL_IMAGE = 6        # F: Image Link (PRIORITY - first to populate)
    COL_TITLE = 7        # G: Description
    COL_PRICE = 8        # H: Price
    COL_URL = 9          # I: Product URL
    COL_DETAIL_IMAGE = 10 # J: Detail Image Link
    COL_PRODUCT_PAGE = 11 # K: Product Page
    COL_SOURCE = 12       # L: Source Page
    
    source_row_count = ws.max_row

    # Add headers in row 1 for columns F onwards
    headers = ["Image Link", "Description", "Price", "Product URL", "Detail Image Link", "Product Page", "Source Page"]
    columns = [COL_IMAGE, COL_TITLE, COL_PRICE, COL_URL, COL_DETAIL_IMAGE, COL_PRODUCT_PAGE, COL_SOURCE]
    for col, header in zip(columns, headers):
        ws.cell(row=1, column=col, value=header)

    print("   Added headers in row 1 for columns F-L")
    print("   Matching catalog numbers from column B (B: Mfr Catalog No.)...")
    print("   Populating from column F onwards (preserving columns A-E)...")
    
    for row_idx in range(2, ws.max_row + 1):
        catalog_cell = ws.cell(row=row_idx, column=COL_CATALOG)
        catalog_number = normalize_catalog_number(catalog_cell.value)
        
        if not catalog_number:
            continue
        
        if catalog_number.lower() in {"mfr catalog no.", "catalog", "catalog no."}:
            continue
        
        # Find matching JSON item
        candidates = catalog_lookup.get(catalog_number, [])
        json_item = pick_best_item(candidates)
        
        if json_item:
            try:
                # F: Image Link (PRIORITY - Column 6)
                image_url = get_preferred_image(json_item)
                ws.cell(row=row_idx, column=COL_IMAGE, value=image_url)
                
                # G: Description (Title)
                title = json_item.get('title', '')
                ws.cell(row=row_idx, column=COL_TITLE, value=title)
                
                # H: Price
                price = get_preferred_price(json_item)
                ws.cell(row=row_idx, column=COL_PRICE, value=price)
                
                # I: Product URL
                url = get_product_url(json_item)
                ws.cell(row=row_idx, column=COL_URL, value=url)
                
                # J: Detail Image Link
                detail_image = get_detail_image(json_item)
                ws.cell(row=row_idx, column=COL_DETAIL_IMAGE, value=detail_image)
                
                # K: Product Page
                product_page = json_item.get('product_page', '')
                ws.cell(row=row_idx, column=COL_PRODUCT_PAGE, value=product_page)
                
                # L: Source Page
                source_page = json_item.get('source_page', '')
                ws.cell(row=row_idx, column=COL_SOURCE, value=source_page)
                
                matched_count += 1
            except Exception as e:
                print(f"[ERR] Error populating row {row_idx}: {e}")
        else:
            unmatched_count += 1
            if unmatched_count <= 5:
                print(f"  [WARN] No match found for catalog: {catalog_number}")

    print(f"   Source row count before populate: {source_row_count}")
    print(f"   Row count after populate:         {ws.max_row}")
    
    return matched_count, unmatched_count

def main():
    print("=" * 70)
    print("ETERNUM Excel Populator - JSON -> ETERNUM.xlsx")
    print("=" * 70)
    print("Mode: Populate from column F onwards only")
    print("      (Columns A-E remain UNTOUCHED)")
    print("=" * 70)
    
    try:
        # Load JSON data
        print("\n1. Loading JSON data...")
        json_data, loaded_sources = load_all_json_data(JSON_FILES)
        for source_path, source_count in loaded_sources:
            print(f"   [OK] Loaded {source_count} products from: {source_path}")
        print(f"   [OK] Total loaded products: {len(json_data)}")
        
        # Load Excel
        print("\n2. Loading Excel file...")
        wb, ws = load_excel()
        
        # Populate Excel
        print("\n3. Populating Excel with JSON data...")
        matched, unmatched = populate_excel(wb, ws, json_data)
        
        # Save Excel
        print("\n4. Saving Excel file...")
        wb.save(EXCEL_FILE)
        print(f"   [OK] Saved: {EXCEL_FILE}")
        
        # Summary
        print("\n" + "=" * 70)
        print("SUMMARY")
        print("=" * 70)
        print(f"[OK] Matched & Populated:  {matched} rows")
        print(f"[WARN] Unmatched Rows:     {unmatched} rows")
        print(f"[OK] Image Link Priority:  Column F (populated first)")
        print(f"[OK] Other Data:           Columns G-L")
        print(f"[OK] Protected Columns:    A-E (unchanged)")
        print("=" * 70)
        
    except FileNotFoundError as e:
        print(f"\n[ERR] File not found - {e}")
        print(f"  Make sure all input files exist:")
        for json_path in JSON_FILES:
            print(f"    - {json_path}")
        print(f"    - {EXCEL_FILE}")
    except json.JSONDecodeError as e:
        print(f"\n[ERR] Invalid JSON - {e}")
    except Exception as e:
        print(f"\n[ERR] Unexpected error: {e}")

if __name__ == "__main__":
    main()
